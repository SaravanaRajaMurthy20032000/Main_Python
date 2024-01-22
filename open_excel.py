from openpyxl import load_workbook

workbook = load_workbook(filename='C:/Users/HP/Downloads/ot.xlsx')

res = workbook.active
for row in range(0, res.max_row):
    for col in res.iter_cols(1, res.max_column):
        print(col[row].value, end=" ")
    print()

# 
import pandas as pd

df = pd.read_excel('C:/Users/HP/Documents/book.xlsx')
print(df)